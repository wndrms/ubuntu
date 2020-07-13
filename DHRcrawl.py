from bs4 import BeautifulSoup
from explicit import waiter, XPATH
from selenium import webdriver
import openpyxl
import sys
import urllib.request
import os
import zipfile

driver_path = '/usr/bin/chromedriver'

def retrieve_file_paths(dirName):
    filePaths = []

    for root, directories, files in os.walk(dirName):
        for filename in files:
            filePath = os.path.join(root, filename)
            filePaths.append(filePath)

    return filePaths

def zip_downloads():
    dir_name = 'downloads'

    filePaths = retrieve_file_paths(dir_name)

    print('The following list of files will be zipped:\n')
    for fileName in filePaths:
        print(fileName)

    zip_file = zipfile.ZipFile(dir_name + '.zip', 'w')
    with zip_file:
        for file in filePaths:
            zip_file.write(file)
    print(dir_name + '.zip file is created successfully!\n')

def removeExtensionFile(filePath, fileExtension):
    if os.path.exists(filePath):
        for file in os.scandir(filePath):
            if file.name.endswith(fileExtension):
                os.remove(file.path)
        return 'Remove File :' + fileExtension
    else:
        return 'Directory Not Found'

def login(driver):
    username = "dlcjs1104"
    password = "slwm2000"

    driver.get("http://danharoo.com/member/login.html")

    waiter.find_write(driver, "//*[@id='member_id']", username, by=XPATH)
    waiter.find_write(driver, "//*[@id='member_passwd']", password, by=XPATH)
    submit = driver.find_element_by_xpath("/html/body/div[4]/div/div/form/div/div/fieldset/a")
    submit.click()

    waiter.find_element(driver, "//*[@id='contents_main']/div[1]/div[1]/ul/li[11]/a/img", by=XPATH)

def crawlDHR_cate(driver, cate_no):
    URL = "http://danharoo.com/product/list.html?cate_no="+cate_no
    try:
        driver.get(URL)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        tag = soup.select(
            'ul>li>div>a'
        )
        product_list = ''
        for x in tag :
            s = str(x).find('product_no=')
            product_list = product_list + str(x)[s+11:s+16] +','

        return product_list

def crawlDHR(driver, product_no, wb, i):
    URL = "http://danharoo.com/product/detail.html?product_no="+product_no
    #wb = openpyxl.load_workbook('/var/www/html/prod_batch_sample.xlsx')
    #wb = openpyxl.load_workbook('prod_batch_sample.xlsx')
    ws = wb.active

    try: 
        driver.get(URL)
        name = driver.find_element_by_class_name("item_name").text
        price = driver.find_element_by_id("span_product_price_text").text
        price = price.replace(',','')
        price = price.rstrip('won')
        price_sell = int(int(price)*1.3)
        price_sell = int((price_sell+90)/100)*100
        price_old = int(int(price)*1.8)
        price_old = int((price_old+90)/100)*100
        product_code = driver.find_element_by_xpath("/html/body/div[4]/div/div/div[2]/div[2]/div[2]/div[4]/table/tbody/tr[8]/td/span").text
        detail = driver.find_element_by_xpath("/html/body/div[4]/div/div/div[2]/div[2]/div[2]/div[4]/table/tbody/tr[6]/td/span").text
        img1 = driver.find_element_by_xpath("/html/body/div[4]/div/div/div[2]/div[2]/div[1]/div[1]/div[1]/a/img").get_attribute('src')
        urllib.request.urlretrieve(img1, "/var/www/html/downloads/sample"+str(i-3)+"-1.jpg")
        img2 = driver.find_element_by_xpath("/html/body/div[4]/div/div/div[2]/div[2]/div[1]/div[2]/ul/li[2]/img").get_attribute('src')
        img2 = img2.replace('small', 'big')
        urllib.request.urlretrieve(img2, "/var/www/html/downloads/sample"+str(i-3)+"-2.jpg")
        img3 = driver.find_element_by_xpath("/html/body/div[4]/div/div/div[2]/div[2]/div[1]/div[2]/ul/li[3]/img").get_attribute('src')
        img3 = img3.replace('small', 'big')
        urllib.request.urlretrieve(img3, "/var/www/html/downloads/sample"+str(i-3)+"-3.jpg") 
        img4 = driver.find_element_by_xpath("/html/body/div[4]/div/div/div[2]/div[2]/div[1]/div[2]/ul/li[4]/img").get_attribute('src')
        img4 = img4.replace('small', 'big')
        urllib.request.urlretrieve(img4, "/var/www/html/downloads/sample"+str(i-3)+"-4.jpg") 
        img5 = driver.find_element_by_xpath("/html/body/div[4]/div/div/div[2]/div[2]/div[1]/div[2]/ul/li[5]/img").get_attribute('src')
        img5 = img5.replace('small', 'big')
        urllib.request.urlretrieve(img5, "/var/www/html/downloads/sample"+str(i-3)+"-5.jpg") 
        tmp_select = driver.find_element_by_xpath("/html/body/div[4]/div/div/div[2]/div[2]/div[2]/div[6]/div[4]/div/div[1]/div/div[1]/table/tbody[2]/tr/td/select/option[3]").click()

        img_list = 'sample' + str(i-3) + '-1.jpg'
        t = 2
        while t < 6 :
            img_list = img_list + ',' + 'sample' + str(i-3) + '-' + str(t) + '.jpg'
            t = t + 1
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        tmp_detail = soup.find('div', class_='cont')
        product_detail = []
        for x in tmp_detail:
            product_detail.append(str(x).replace('ec-data-src', 'src'))

        select1 = soup.find('select', id='product_option_id1')
        select2 = soup.find('select', id='product_option_id2')
        select1 = select1.select(
                'option'
        )
        if len(select1) > 2 :
            select_list = ''
            t = 3
            select_list = select_list + select1[2].text
            while t < len(select1) :
                select_list = select_list + ',' + select1[t].text
                t = t + 1
            select_list = select_list + '\n'
            select2 = select2.select(
                    'option'
            )
            t = 3
            select_list = select_list + select2[2].text
            while t < len(select2) :
                select_list = select_list + ',' + select2[t].text
                t = t + 1
            price_list = str(price_sell)
            t = 1
            while t < (len(select1)-2) :
                price_list = price_list + ',' + str(price_sell)
                t = t + 1

            ws.cell(row=i, column=1).value = name
            ws.cell(row=i, column=2).value = product_code
            ws.cell(row=i, column=3).value = 'CATE66'
            ws.cell(row=i, column=4).value = detail
            ws.cell(row=i, column=7).value = price_sell
            ws.cell(row=i, column=8).value = 1
            ws.cell(row=i, column=9).value = price_old
            ws.cell(row=i, column=10).value = 'N'
            ws.cell(row=i, column=13).value = img_list
            ws.cell(row=i, column=14).value = product_detail[1]
            ws.cell(row=i, column=15).value = '과세상품'
            ws.cell(row=i, column=16).value = 'Y'
            ws.cell(row=i, column=17).value = 'N'
            ws.cell(row=i, column=21).value = '택배/소포/등기'
            ws.cell(row=i, column=22).value = 'Y'
            ws.cell(row=i, column=23).value = 'N'
            ws.cell(row=i, column=24).value = 'N'
            ws.cell(row=i, column=25).value = 'P'
            ws.cell(row=i, column=41).value = 'Y'
            ws.cell(row=i, column=42).value = 'N'
            ws.cell(row=i, column=43).value = 0
            ws.cell(row=i, column=44).value = '원'
            ws.cell(row=i, column=45).value = '조합형'
            ws.cell(row=i, column=46).value = 'Color\nSize'
            ws.cell(row=i, column=47).value = select_list
            ws.cell(row=i, column=48).value = price_list
            ws.cell(row=i, column=56).value = 'N'
            ws.cell(row=i, column=63).value = 'N'
            ws.cell(row=i, column=64).value = 'N'
            ws.cell(row=i, column=65).value = 'N'
            ws.cell(row=i, column=67).value = 'N'
            ws.cell(row=i, column=68).value = 'N'
            ws.cell(row=i, column=69).value = 'N'
            ws.cell(row=i, column=70).value = 'N'
        else :
            print(product_no + " sold out")
    except:
        print(product_no + " 제품이 존재하지 않습니다")

if __name__ == '__main__':
    product_no = sys.argv[1].split(',')
    i=4
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')

    driver = webdriver.Chrome(driver_path, chrome_options=options)
    wb = openpyxl.load_workbook('/var/www/html/prod_batch_sample.xlsx')

    try:
        login(driver)
        for x in product_no:
            crawlDHR(driver, x, wb, i)
            i=i+1
    finally:
        wb.remove(wb['시트1'])
        wb.save('/var/www/html/downloads/batch.xlsx')
        wb.close()
        driver.quit()
        print("crawl Down\n")
        zip_downloads()
        print(removeExtensionFile('/var/www/html/downloads', '.jpg'))

